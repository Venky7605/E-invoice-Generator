"""
Excel and ZIP export utilities for bulk download.
"""

import io
import zipfile
from datetime import date

import openpyxl
from utils.nic_api import STATE_NAMES
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


_HEADER_FILL  = PatternFill("solid", fgColor="1A3C5E")
_HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
_ALT_FILL     = PatternFill("solid", fgColor="F0F4F8")
_BORDER_SIDE  = Side(style="thin", color="CBD5E1")
_BORDER       = Border(left=_BORDER_SIDE, right=_BORDER_SIDE,
                       top=_BORDER_SIDE, bottom=_BORDER_SIDE)
_CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT         = Alignment(horizontal="left",  vertical="center", wrap_text=True)
_RIGHT        = Alignment(horizontal="right", vertical="center")


def _style_header(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER


def _auto_width(ws, min_w=8, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def export_invoices_excel(invoices: dict) -> bytes:
    """Export all invoices to a multi-sheet Excel workbook."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Invoice Register ────────────────────────────────────
    ws = wb.active
    ws.title = "Invoice Register"
    ws.freeze_panes = "A2"

    headers = [
        "Status", "Doc Type", "Invoice No", "Date", "Supply Type",
        "Buyer Name", "Buyer GSTIN", "State",
        "Taxable (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)",
        "Other Charges (₹)", "Round Off (₹)", "Total Invoice (₹)",
        "IRN", "Ack No", "Ack Date", "Mode",
    ]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.row_dimensions[1].height = 22

    for row_idx, (k, inv) in enumerate(sorted(
            invoices.items(),
            key=lambda x: x[1].get("_saved_at", ""), reverse=True), start=2):
        doc   = inv.get("DocDtls", {})
        val   = inv.get("ValDtls", {})
        buy   = inv.get("BuyerDtls", {})
        tran  = inv.get("TranDtls", {})
        irnd  = inv.get("_irn_data", {})
        st    = inv.get("_status", "PENDING")

        state = STATE_NAMES.get(buy.get("Stcd", ""), "")

        row_data = [
            st,
            doc.get("Typ", ""),
            doc.get("No", ""),
            doc.get("Dt", ""),
            tran.get("SupTyp", ""),
            buy.get("LglNm", ""),
            buy.get("Gstin", ""),
            state,
            float(val.get("AssVal", 0)),
            float(val.get("IgstVal", 0)),
            float(val.get("CgstVal", 0)),
            float(val.get("SgstVal", 0)),
            float(val.get("CesVal", 0)),
            float(val.get("OthChrg", 0)),
            float(val.get("RndOffAmt", 0)),
            float(val.get("TotInvVal", 0)),
            irnd.get("irn", ""),
            irnd.get("ack_no", ""),
            irnd.get("ack_dt", ""),
            "Test" if irnd.get("simulated") else ("Live" if irnd else "—"),
        ]
        ws.append(row_data)
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        for c in range(1, len(headers) + 1):
            cell            = ws.cell(row=row_idx, column=c)
            cell.border     = _BORDER
            if fill:
                cell.fill   = fill
            if c in (9, 10, 11, 12, 13, 14, 15, 16):
                cell.number_format = '₹#,##0.00'
                cell.alignment     = _RIGHT
            else:
                cell.alignment     = _LEFT

    _auto_width(ws)

    # ── Sheet 2: Line Items ──────────────────────────────────────────
    ws2 = wb.create_sheet("Line Items")
    ws2.freeze_panes = "A2"

    item_hdrs = [
        "Invoice No", "Date", "Buyer", "Sl No", "Description",
        "HSN/SAC", "Is Service", "Qty", "UQC", "Unit Price (₹)",
        "Discount (₹)", "Taxable (₹)", "GST Rate (%)",
        "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)", "Item Total (₹)",
    ]
    ws2.append(item_hdrs)
    _style_header(ws2, 1, len(item_hdrs))

    row_idx = 2
    for k, inv in invoices.items():
        doc   = inv.get("DocDtls", {})
        buy   = inv.get("BuyerDtls", {})
        items = inv.get("ItemList", [])
        for item in items:
            ws2.append([
                doc.get("No", ""),
                doc.get("Dt", ""),
                buy.get("LglNm", ""),
                item.get("SlNo", ""),
                item.get("PrdDesc", ""),
                item.get("HsnCd", ""),
                item.get("IsServc", "N"),
                float(item.get("Qty", 0)),
                item.get("Unit", ""),
                float(item.get("UnitPrice", 0)),
                float(item.get("Discount", 0)),
                float(item.get("AssAmt", 0)),
                float(item.get("GstRt", 0)),
                float(item.get("IgstAmt", 0)),
                float(item.get("CgstAmt", 0)),
                float(item.get("SgstAmt", 0)),
                float(item.get("CesAmt", 0)),
                float(item.get("TotItemVal", 0)),
            ])
            for c in range(1, len(item_hdrs) + 1):
                cell        = ws2.cell(row=row_idx, column=c)
                cell.border = _BORDER
                if c in (8, 10, 11, 12, 14, 15, 16, 17, 18):
                    cell.number_format = '₹#,##0.00'
                    cell.alignment     = _RIGHT
            row_idx += 1

    _auto_width(ws2)

    # ── Sheet 3: HSN Summary ─────────────────────────────────────────
    ws3 = wb.create_sheet("HSN Summary")
    hsn_map = {}
    for inv in invoices.values():
        for item in inv.get("ItemList", []):
            hsn  = item.get("HsnCd", "UNK")
            desc = item.get("PrdDesc", "")
            uqc  = item.get("Unit", "OTH")
            if hsn not in hsn_map:
                hsn_map[hsn] = {"desc": desc, "uqc": uqc, "qty": 0,
                                "val": 0, "txval": 0, "igst": 0,
                                "cgst": 0, "sgst": 0, "cess": 0}
            hsn_map[hsn]["qty"]   += float(item.get("Qty", 0))
            hsn_map[hsn]["val"]   += float(item.get("TotItemVal", 0))
            hsn_map[hsn]["txval"] += float(item.get("AssAmt", 0))
            hsn_map[hsn]["igst"]  += float(item.get("IgstAmt", 0))
            hsn_map[hsn]["cgst"]  += float(item.get("CgstAmt", 0))
            hsn_map[hsn]["sgst"]  += float(item.get("SgstAmt", 0))
            hsn_map[hsn]["cess"]  += float(item.get("CesAmt", 0))

    hsn_hdrs = ["HSN/SAC", "Description", "UQC", "Qty",
                "Total Value (₹)", "Taxable (₹)", "IGST (₹)",
                "CGST (₹)", "SGST (₹)", "Cess (₹)"]
    ws3.append(hsn_hdrs)
    _style_header(ws3, 1, len(hsn_hdrs))
    for i, (hsn, v) in enumerate(hsn_map.items(), start=2):
        ws3.append([hsn, v["desc"], v["uqc"], round(v["qty"], 3),
                    round(v["val"], 2), round(v["txval"], 2),
                    round(v["igst"], 2), round(v["cgst"], 2),
                    round(v["sgst"], 2), round(v["cess"], 2)])
        for c in range(1, len(hsn_hdrs) + 1):
            ws3.cell(row=i, column=c).border = _BORDER
    _auto_width(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def export_pdfs_zip(invoices: dict, pdf_fn) -> bytes:
    """
    Build a ZIP of e-invoice PDFs.
    pdf_fn(invoice, irn_data) → BytesIO  (generate_einvoice_pdf)
    """
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for inv_key, inv in invoices.items():
            irnd = inv.get("_irn_data", {})
            doc  = inv.get("DocDtls", {})
            try:
                pdf_io = pdf_fn(inv, irnd)
                fname  = f"einvoice_{doc.get('No', inv_key).replace('/', '_')}.pdf"
                zf.writestr(fname, pdf_io.getvalue())
            except Exception:
                pass
    buf.seek(0)
    return buf.read()
